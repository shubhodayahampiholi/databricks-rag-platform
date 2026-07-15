"""
Extraction implementations. Each function takes raw bytes and returns a
common ExtractionResult, so downstream chunking can work uniformly
regardless of which method produced the text.

Sections preserve structure relevant to the source type -- pages for PDF,
heading-delimited blocks for DOCX, individual rows for XLSX -- because
chunking strategies (heading_hierarchy_split, table_aware_row_based) need
that structure, not just a flat string. See docs/DESIGN.md, Silver Layer.
"""
import io
from dataclasses import dataclass, field
from typing import Optional

import fitz  # pymupdf
from docx import Document as DocxDocument
from openpyxl import load_workbook


@dataclass
class ExtractionSection:
    heading: Optional[str]
    text: str
    position: int  # page number, row index, or paragraph block index


@dataclass
class ExtractionResult:
    status: str  # "success" | "empty" | "corrupt"
    full_text: str
    sections: list[ExtractionSection] = field(default_factory=list)
    error_message: Optional[str] = None


def pymupdf_standard(content: bytes) -> ExtractionResult:
    """
    Standard text extraction for born-digital PDFs. One section per page.
    Does not attempt OCR -- a scanned/image-only PDF will correctly
    produce empty or near-empty text here, which is the expected,
    honest result for this method, not a bug in it.
    """
    try:
        doc = fitz.open(stream=content, filetype="pdf")
    except Exception as e:
        return ExtractionResult(status="corrupt", full_text="", error_message=str(e))

    sections = []
    full_text_parts = []
    for page_num, page in enumerate(doc):
        page_text = page.get_text().strip()
        sections.append(
            ExtractionSection(heading=None, text=page_text, position=page_num)
        )
        full_text_parts.append(page_text)
    doc.close()

    full_text = "\n\n".join(full_text_parts).strip()
    status = "success" if full_text else "empty"
    return ExtractionResult(status=status, full_text=full_text, sections=sections)


def python_docx_standard(content: bytes) -> ExtractionResult:
    """
    DOCX extraction that preserves heading structure. Paragraphs are
    grouped into sections under their most recent heading, so
    heading_hierarchy_split has real structure to chunk against rather
    than a flat blob of text.
    """
    try:
        doc = DocxDocument(io.BytesIO(content))
    except Exception as e:
        return ExtractionResult(status="corrupt", full_text="", error_message=str(e))

    sections = []
    current_heading = None
    current_text_parts = []
    position = 0
    full_text_parts = []

    def flush_section():
        nonlocal current_text_parts, position
        if current_text_parts:
            text = "\n".join(current_text_parts).strip()
            if text:
                sections.append(
                    ExtractionSection(
                        heading=current_heading, text=text, position=position
                    )
                )
                position += 1
            current_text_parts = []

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        full_text_parts.append(text)
        is_heading = para.style.name.startswith("Heading")
        if is_heading:
            flush_section()
            current_heading = text
        else:
            current_text_parts.append(text)

    flush_section()

    full_text = "\n\n".join(full_text_parts).strip()
    status = "success" if full_text else "empty"
    return ExtractionResult(status=status, full_text=full_text, sections=sections)


def openpyxl_standard(content: bytes) -> ExtractionResult:
    """
    XLSX extraction with one section per row, not a flattened blob --
    table_aware_row_based chunking needs individual row boundaries, not
    a single string representing the whole sheet.
    """
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        return ExtractionResult(status="corrupt", full_text="", error_message=str(e))

    sections = []
    full_text_parts = []
    position = 0

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c) if c is not None else "" for c in rows[0]]
        for row in rows[1:]:
            row_values = [str(c) if c is not None else "" for c in row]
            row_text = ", ".join(
                f"{h}: {v}" for h, v in zip(header, row_values) if v
            )
            if not row_text:
                continue
            sections.append(
                ExtractionSection(
                    heading=sheet.title, text=row_text, position=position
                )
            )
            full_text_parts.append(row_text)
            position += 1

    full_text = "\n".join(full_text_parts).strip()
    status = "success" if full_text else "empty"
    return ExtractionResult(status=status, full_text=full_text, sections=sections)


def plain_text_decode(content: bytes) -> ExtractionResult:
    """
    Markdown/plain text: no library needed, just a UTF-8 decode. One
    section, since there's no structure to preserve beyond the raw text
    itself -- markdown heading parsing is a reasonable future refinement,
    not attempted here.
    """
    try:
        text = content.decode("utf-8").strip()
    except UnicodeDecodeError as e:
        return ExtractionResult(status="corrupt", full_text="", error_message=str(e))

    status = "success" if text else "empty"
    sections = [ExtractionSection(heading=None, text=text, position=0)] if text else []
    return ExtractionResult(status=status, full_text=text, sections=sections)